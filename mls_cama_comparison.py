from IPython import get_ipython
from IPython.display import display
import pandas as pd
import numpy as np
import os

# Install required package for Excel hyperlinks if not available
try:
    import openpyxl
except ImportError:
    import subprocess
    import sys
    print("Installing openpyxl for Excel file creation...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--break-system-packages", "openpyxl"])
    import openpyxl

# --- Configuration ---
MLS_DATA_PATH = '/MLS_11-7-25.xlsx'
CAMA_DATA_PATH = '/CAMA_OCT_31.xls'

UNIQUE_ID_COLUMN = {'mls_col': 'Parcel Number', 'cama_col': 'PARID'}

# CORRECTED column mappings based on actual Excel file headers
COLUMNS_TO_COMPARE = [
    {'mls_col': 'Above Grade Finished Area', 'cama_col': 'SFLA'},
    {'mls_col': 'Bedrooms Total', 'cama_col': 'RMBED'},
    {'mls_col': 'Bathrooms Full', 'cama_col': 'FIXBATH'},
    {'mls_col': 'Bathrooms Half', 'cama_col': 'FIXHALF'},
]

# SPECIAL COMPARISON: Sum multiple CAMA columns to compare against one MLS column
# MLS "Below Grade Finished Area" should equal sum of CAMA (RECROMAREA + FINBSMTAREA + UFEATAREA)
COLUMNS_TO_COMPARE_SUM = [
    {'mls_col': 'Below Grade Finished Area', 'cama_cols': ['RECROMAREA', 'FINBSMTAREA', 'UFEATAREA']}
]

# CATEGORICAL COMPARISON: Check if MLS field contains text, map to CAMA numeric value
# If MLS "Cooling" contains "Central Air", CAMA "HEAT" should be 1; otherwise 0
COLUMNS_TO_COMPARE_CATEGORICAL = [
    {
        'mls_col': 'Cooling',
        'cama_col': 'HEAT',
        'mls_check_contains': 'Central Air',  # Check if MLS contains this text
        'cama_expected_if_true': 1,            # Expected CAMA value if MLS contains text
        'cama_expected_if_false': 0,           # Expected CAMA value if MLS does not contain text
        'case_sensitive': False                # Case-insensitive search
    }
]

# TOLERANCE SETTING - Adjust this if needed!
NUMERIC_TOLERANCE = 0.01  # Absolute tolerance for numeric comparisons

# SKIP ZERO VALUES - Set to True if 0 in MLS means "no data"
SKIP_ZERO_VALUES = True  # Change to False if 0 is a valid value to compare

# ==================================================================================
# HYPERLINK CONFIGURATION - WINDOW ID INPUT
# ==================================================================================
print("=" * 80)
print("MLS vs. CAMA Data Comparison Tool")
print("=" * 80)
print()
print("📌 How to get WindowId:")
print("   1. Go to https://iasworld.starkcountyohio.gov/iasworld/")
print("   2. Log in and search for any property")
print("   3. Look at the URL and copy the windowId value")
print("   Example: ...windowId=638981240146803746&...")
print()

# Default windowId
DEFAULT_WINDOW_ID = "638981240146803746"

# Ask user for windowId
user_input = input(f"Enter WindowId (or press Enter to use default: {DEFAULT_WINDOW_ID}): ").strip()

if user_input:
    WINDOW_ID = user_input
    print(f"✅ Using your windowId: {WINDOW_ID}")
else:
    WINDOW_ID = DEFAULT_WINDOW_ID
    print(f"✅ Using default windowId: {WINDOW_ID}")

print()

PARCEL_ID_URL_TEMPLATE = f"https://iasworld.starkcountyohio.gov/iasworld/Maintain/Transact.aspx?txtMaskedPin={{parcel_id}}&selYear=&userYear=&selJur=&chkShowHistory=False&chkShowChanges=&chkShowDeactivated=&PinValue={{parcel_id}}&pin=&trans_key=&windowId={WINDOW_ID}&submitFlag=true&TransPopUp=&ACflag=False&ACflag2=False"

# Zillow URL - will use search format since we don't have zpid
ZILLOW_URL_BASE = "https://www.zillow.com/homes/"

# MLS column names for address components (update these if your column names are different)
ADDRESS_COLUMNS = {
    'address': 'Address',  # Street address
    'city': 'City',
    'state': 'State or Province',
    'zip': 'Postal Code'
}

def format_zillow_url(address, city, state, zip_code):
    """
    Create a Zillow search URL from address components.
    Example: https://www.zillow.com/homes/1610-20th-St-NW-Canton-OH-44709_rb/
    """
    if pd.isna(address) or pd.isna(city) or pd.isna(zip_code):
        return None
    
    # Clean and format each component
    # Replace spaces with hyphens, remove special characters
    import re
    
    address_clean = str(address).strip()
    city_clean = str(city).strip()
    zip_clean = str(zip_code).strip().split('-')[0]  # Remove ZIP+4 if present
    
    # Remove apartment/unit numbers (e.g., "Apt 2", "Unit B", "#3")
    address_clean = re.sub(r'\s+(Apt|Unit|#|Suite)\s*[\w-]*$', '', address_clean, flags=re.IGNORECASE)
    
    # Replace spaces with hyphens and remove special characters except hyphens
    address_formatted = re.sub(r'[^\w\s-]', '', address_clean)
    address_formatted = re.sub(r'\s+', '-', address_formatted)
    
    city_formatted = re.sub(r'[^\w\s-]', '', city_clean)
    city_formatted = re.sub(r'\s+', '-', city_formatted)
    
    # Construct the Zillow search URL - format: address-city-OH-zip_rb/
    url_slug = f"{address_formatted}-{city_formatted}-OH-{zip_clean}_rb"
    
    return f"{ZILLOW_URL_BASE}{url_slug}/"

# --- Data Loading Functions ---

def read_mls_data(file_path):
    """Reads MLS data from a specified Excel file."""
    try:
        df_mls = pd.read_excel(file_path)
        print(f"Successfully loaded MLS data from: {file_path}")
        return df_mls
    except FileNotFoundError:
        print(f"Error: MLS data file not found at {file_path}")
        return None
    except Exception as e:
        print(f"Error reading MLS data: {e}")
        return None

def read_cama_data(file_path):
    """Reads CAMA system data from a specified Excel file."""
    try:
        df_cama = pd.read_excel(file_path)
        print(f"Successfully loaded CAMA data from: {file_path}")
        return df_cama
    except FileNotFoundError:
        print(f"Error: CAMA data file not found at {file_path}")
        return None
    except Exception as e:
        print(f"Error reading CAMA data: {e}")
        return None

# --- Data Analysis Functions ---

def find_duplicate_ids(df, id_column, source_name):
    """Finds and reports duplicate IDs within a single DataFrame."""
    if df is None or df.empty:
        print(f"No data to check for duplicates in {source_name}.")
        return pd.DataFrame()

    duplicate_ids = df[df.duplicated(subset=[id_column], keep=False)]

    if not duplicate_ids.empty:
        print(f"\n--- Duplicate '{id_column}'s found within {source_name} data ---")
        print(duplicate_ids.sort_values(by=id_column).to_string())
        return duplicate_ids
    else:
        print(f"\nNo duplicate '{id_column}'s found within {source_name} data.")
        return pd.DataFrame()

def values_equal(val1, val2):
    """Helper function to check if two values are equal using the same logic as main comparison."""
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')

        if pd.isna(num1) and pd.isna(num2):
            return True
        elif pd.isna(num1) != pd.isna(num2):
            return False
        else:
            return np.isclose(num1, num2, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except (ValueError, TypeError):
        str1 = str(val1).strip().lower() if pd.notna(val1) else ''
        str2 = str(val2).strip().lower() if pd.notna(val2) else ''
        return str1 == str2

def categorical_match(mls_val, cama_val, mapping):
    """
    Check if a categorical MLS field matches expected CAMA value based on text content.

    Args:
        mls_val: MLS field value to check for text
        cama_val: CAMA field value to validate
        mapping: Dict with comparison rules

    Returns:
        True if values match expected mapping, False otherwise
    """
    # Extract mapping rules
    check_text = mapping.get('mls_check_contains', '')
    expected_if_true = mapping.get('cama_expected_if_true')
    expected_if_false = mapping.get('cama_expected_if_false')
    case_sensitive = mapping.get('case_sensitive', False)

    # Convert MLS value to string for text searching
    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''

    # Apply case sensitivity
    if not case_sensitive:
        mls_str = mls_str.lower()
        check_text = check_text.lower()

    # Check if text is found in MLS field
    text_found = check_text in mls_str

    # Determine expected CAMA value based on text match
    expected_cama = expected_if_true if text_found else expected_if_false

    # Compare with actual CAMA value
    try:
        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
        expected_numeric = pd.to_numeric(expected_cama, errors='coerce')

        if pd.isna(cama_numeric) and pd.isna(expected_numeric):
            return True
        elif pd.isna(cama_numeric) or pd.isna(expected_numeric):
            return False
        else:
            return np.isclose(cama_numeric, expected_numeric, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except:
        return str(cama_val).strip().lower() == str(expected_cama).strip().lower()

# --- Enhanced Data Comparison Function ---

def compare_data_enhanced(df_mls, df_cama, unique_id_col, cols_to_compare_mapping,
                         cols_to_compare_sum=None, cols_to_compare_categorical=None, debug_mode=False):
    """
    Compares MLS and CAMA dataframes with enhanced mismatch reporting.
    Returns separate DataFrames for different discrepancy types AND perfect matches.

    Args:
        df_mls: MLS DataFrame
        df_cama: CAMA DataFrame
        unique_id_col: Dict with 'mls_col' and 'cama_col' keys
        cols_to_compare_mapping: List of dicts for 1-to-1 column comparisons
        cols_to_compare_sum: List of dicts with 'mls_col' and 'cama_cols' (list) for sum comparisons
        cols_to_compare_categorical: List of dicts for categorical comparisons
        debug_mode: Boolean for debug output
    """
    if df_mls is None or df_cama is None:
        print("Cannot compare data: one or both dataframes are missing.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if not cols_to_compare_mapping:
        print("Cannot compare data: Column mapping is empty or invalid.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    mls_id_col_name = unique_id_col.get('mls_col')
    cama_id_col_name = unique_id_col.get('cama_col')

    if mls_id_col_name is None or cama_id_col_name is None:
        print("Error: Unique ID column mapping is incomplete or invalid.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Rename MLS unique ID column to match CAMA for merging
    if mls_id_col_name in df_mls.columns:
        df_mls_renamed = df_mls.copy()
        df_mls_renamed = df_mls_renamed.rename(columns={mls_id_col_name: cama_id_col_name})
    else:
        print(f"Error: Unique ID column '{mls_id_col_name}' not found in MLS data.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    if cama_id_col_name not in df_cama.columns:
        print(f"Error: Unique ID column '{cama_id_col_name}' not found in CAMA data.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Perform merges - NOTE: No overlapping column names means NO SUFFIXES are added!
    matched_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='inner')
    merged_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='outer', indicator=True)

    # Lists to store different types of discrepancies
    missing_in_cama = []
    missing_in_mls = []
    value_mismatches = []
    perfect_matches = []

    comparison_debug = []

    # Iterate through the merged DataFrame
    for index, row in merged_df.iterrows():
        record_id = row.get(cama_id_col_name)
        merge_status = row.get('_merge')

        if merge_status == 'left_only':
            # Extract Listing # and Closed Date from MLS data
            listing_num = row.get('Listing #', '')
            closed_date = row.get('Closed Date', '')

            missing_in_cama.append({
                'Parcel_ID': record_id,
                'Listing_Number': listing_num,
                'Closed_Date': closed_date
            })

        elif merge_status == 'right_only':
            missing_in_mls.append({'Parcel_ID': record_id})

        elif merge_status == 'both':
            # Extract Listing #, SALEKEY, NOPAR, and ADDITIONAL_PARCELS once for this record
            listing_num = row.get('Listing #', '')
            salekey = row.get('SALEKEY', '')
            nopar = row.get('NOPAR', '')
            additional_parcels = row.get('ADDITIONAL_PARCELS', '')
            
            # Extract address components for Zillow links
            address = row.get(ADDRESS_COLUMNS.get('address', 'Address'), '')
            city = row.get(ADDRESS_COLUMNS.get('city', 'City'), '')
            state = row.get(ADDRESS_COLUMNS.get('state', 'State or Province'), '')
            zip_code = row.get(ADDRESS_COLUMNS.get('zip', 'Postal Code'), '')
            
            record_mismatches = []
            fields_compared = []

            # Standard 1-to-1 comparisons
            for mapping in cols_to_compare_mapping:
                mls_original_col = mapping['mls_col']
                cama_original_col = mapping['cama_col']

                mls_val_col = mls_original_col
                cama_val_col = cama_original_col

                if mls_val_col not in merged_df.columns or cama_val_col not in merged_df.columns:
                    if debug_mode:
                        print(f"⚠ Column not found in merged data: {mls_val_col} or {cama_val_col}")
                    continue

                mls_val = row.get(mls_val_col)
                cama_val = row.get(cama_val_col)

                # SKIP comparison if EITHER value is blank/null/NaN
                mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')

                if mls_is_blank or cama_is_blank:
                    continue

                # Track that we compared this field
                fields_compared.append(mls_original_col)

                # SKIP comparison if EITHER value is 0 and SKIP_ZERO_VALUES is enabled
                if SKIP_ZERO_VALUES:
                    try:
                        mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
                        if (pd.notna(mls_numeric) and mls_numeric == 0) or (pd.notna(cama_numeric) and cama_numeric == 0):
                            continue
                    except:
                        pass

                is_different = not values_equal(mls_val, cama_val)

                if debug_mode:
                    comparison_debug.append({
                        'Parcel_ID': record_id,
                        'Field': mls_original_col,
                        'MLS_Value': mls_val,
                        'CAMA_Value': cama_val,
                        'Is_Different': is_different,
                        'MLS_Type': type(mls_val).__name__,
                        'CAMA_Type': type(cama_val).__name__
                    })

                if is_different:
                    record_mismatches.append({
                        'Parcel_ID': record_id,
                        'NOPAR': nopar,
                        'ADDITIONAL_PARCELS': additional_parcels,
                        'Listing_Number': listing_num,
                        'SALEKEY': salekey,
                        'Address': address,
                        'City': city,
                        'State': state,
                        'Zip': zip_code,
                        'Field_MLS': mls_original_col,
                        'Field_CAMA': cama_original_col,
                        'MLS_Value': mls_val,
                        'CAMA_Value': cama_val,
                        'Difference': calculate_difference(mls_val, cama_val),
                        'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                    })

            # Handle sum comparisons (multiple CAMA columns summed)
            if cols_to_compare_sum:
                for mapping in cols_to_compare_sum:
                    mls_col = mapping['mls_col']
                    cama_cols = mapping['cama_cols']

                    # Check if MLS column exists
                    if mls_col not in merged_df.columns:
                        if debug_mode:
                            print(f"⚠ MLS column not found: {mls_col}")
                        continue

                    # Check if all CAMA columns exist
                    missing_cols = [col for col in cama_cols if col not in merged_df.columns]
                    if missing_cols:
                        if debug_mode:
                            print(f"⚠ CAMA columns not found: {missing_cols}")
                        continue

                    mls_val = row.get(mls_col)

                    # Skip if MLS value is blank
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    if mls_is_blank:
                        continue

                    # Calculate sum of CAMA columns (treating NaN as 0)
                    cama_sum = 0
                    all_cama_blank = True
                    for col in cama_cols:
                        val = row.get(col)
                        if pd.notna(val):
                            all_cama_blank = False
                            try:
                                cama_sum += pd.to_numeric(val, errors='coerce')
                            except:
                                pass

                    # Skip if all CAMA columns are blank
                    if all_cama_blank:
                        continue

                    # Track that we compared this field
                    fields_compared.append(mls_col)

                    # Skip if SKIP_ZERO_VALUES enabled and either side is 0
                    if SKIP_ZERO_VALUES:
                        try:
                            mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                            if (pd.notna(mls_numeric) and mls_numeric == 0) or cama_sum == 0:
                                continue
                        except:
                            pass

                    # Compare MLS value to CAMA sum
                    is_different = not values_equal(mls_val, cama_sum)

                    if debug_mode:
                        comparison_debug.append({
                            'Parcel_ID': record_id,
                            'Field': mls_col,
                            'MLS_Value': mls_val,
                            'CAMA_Value': f"SUM({','.join(cama_cols)})={cama_sum}",
                            'Is_Different': is_different,
                            'MLS_Type': type(mls_val).__name__,
                            'CAMA_Type': 'float (sum)'
                        })

                    if is_different:
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'NOPAR': nopar,
                            'ADDITIONAL_PARCELS': additional_parcels,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Field_MLS': mls_col,
                            'Field_CAMA': f"SUM({', '.join(cama_cols)})",
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_sum,
                            'Difference': calculate_difference(mls_val, cama_sum),
                            'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                        })

            # Handle categorical comparisons
            if cols_to_compare_categorical:
                for mapping in cols_to_compare_categorical:
                    mls_col = mapping['mls_col']
                    cama_col = mapping['cama_col']

                    # Check if columns exist
                    if mls_col not in merged_df.columns:
                        if debug_mode:
                            print(f"⚠ MLS column not found: {mls_col}")
                        continue

                    if cama_col not in merged_df.columns:
                        if debug_mode:
                            print(f"⚠ CAMA column not found: {cama_col}")
                        continue

                    mls_val = row.get(mls_col)
                    cama_val = row.get(cama_col)

                    # Skip if MLS value is blank
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    if mls_is_blank:
                        continue

                    # Skip if CAMA value is blank
                    cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')
                    if cama_is_blank:
                        continue

                    # Track that we compared this field
                    fields_compared.append(mls_col)

                    # Perform categorical comparison
                    is_match = categorical_match(mls_val, cama_val, mapping)

                    # Determine expected value for reporting
                    check_text = mapping.get('mls_check_contains', '')
                    case_sensitive = mapping.get('case_sensitive', False)
                    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''
                    if not case_sensitive:
                        mls_str = mls_str.lower()
                        check_text_lower = check_text.lower()
                    else:
                        check_text_lower = check_text
                    text_found = check_text_lower in mls_str
                    expected_cama = mapping.get('cama_expected_if_true') if text_found else mapping.get('cama_expected_if_false')

                    if debug_mode:
                        comparison_debug.append({
                            'Parcel_ID': record_id,
                            'Field': mls_col,
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_val,
                            'Is_Different': not is_match,
                            'MLS_Type': type(mls_val).__name__,
                            'CAMA_Type': type(cama_val).__name__,
                            'Expected_CAMA': expected_cama
                        })

                    if not is_match:
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'NOPAR': nopar,
                            'ADDITIONAL_PARCELS': additional_parcels,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Field_MLS': mls_col,
                            'Field_CAMA': cama_col,
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_val,
                            'Expected_CAMA_Value': expected_cama,
                            'Match_Rule': f"If '{check_text}' in {mls_col}, then {cama_col} should be {mapping.get('cama_expected_if_true')}, else {mapping.get('cama_expected_if_false')}",
                            'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                        })

            # If no mismatches found for this record, it's a perfect match!
            if not record_mismatches and fields_compared:
                perfect_matches.append({
                    'Parcel_ID': record_id,
                    'NOPAR': nopar,
                    'ADDITIONAL_PARCELS': additional_parcels,
                    'Listing_Number': listing_num,
                    'SALEKEY': salekey,
                    'Address': address,
                    'City': city,
                    'State': state,
                    'Zip': zip_code,
                    'Fields_Compared': len(fields_compared),
                    'Fields_List': ', '.join(fields_compared),
                    'Zillow_URL': format_zillow_url(address, city, state, zip_code)
                })

            value_mismatches.extend(record_mismatches)

    # Convert to DataFrames
    df_missing_cama = pd.DataFrame(missing_in_cama)
    df_missing_mls = pd.DataFrame(missing_in_mls)
    df_value_mismatches = pd.DataFrame(value_mismatches)
    df_perfect_matches = pd.DataFrame(perfect_matches)

    if debug_mode and comparison_debug:
        print(f"\n🔍 DEBUG: Total comparisons made: {len(comparison_debug)}")
        print(f"🔍 DEBUG: Mismatches detected: {sum(1 for c in comparison_debug if c['Is_Different'])}")

        df_debug = pd.DataFrame(comparison_debug).head(20)
        print("\n🔍 DEBUG: First 20 comparisons:")
        display(df_debug)

    return df_missing_cama, df_missing_mls, df_value_mismatches, matched_df, df_perfect_matches

def calculate_difference(val1, val2):
    """Calculate the difference between two values (numeric or text)."""
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')

        if pd.isna(num1) or pd.isna(num2):
            return "N/A"

        diff = num1 - num2
        return f"{diff:,.2f}"
    except (ValueError, TypeError):
        return "Text difference"

# --- Enhanced Reporting Function ---

def report_discrepancies_enhanced(df_missing_cama, df_missing_mls, df_value_mismatches,
                                  df_perfect_matches, output_prefix='discrepancies'):
    """Generates separate reports for each type of discrepancy AND perfect matches with hyperlinks."""
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    
    reports_generated = []

    if not df_missing_cama.empty:
        filename = f"{output_prefix}_missing_in_CAMA.xlsx"
        df_output = df_missing_cama.copy()
        
        # Save to Excel first
        df_output.to_excel(filename, index=False, sheet_name='Missing in CAMA', engine='openpyxl')
        
        # Add hyperlinks using openpyxl
        if 'Parcel_ID' in df_output.columns and PARCEL_ID_URL_TEMPLATE:
            wb = load_workbook(filename)
            ws = wb['Missing in CAMA']
            
            # Find Parcel_ID column
            parcel_col_idx = list(df_output.columns).index('Parcel_ID') + 1
            
            # Add hyperlinks to each cell (skip header row)
            for row_idx in range(2, len(df_output) + 2):
                cell = ws.cell(row=row_idx, column=parcel_col_idx)
                parcel_value = cell.value
                if parcel_value and str(parcel_value).strip():
                    url = PARCEL_ID_URL_TEMPLATE.format(parcel_id=parcel_value)
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'
            
            wb.save(filename)
        
        print(f"\n✓ Missing in CAMA report saved: {filename} ({len(df_missing_cama)} records)")
        reports_generated.append(filename)

    if not df_missing_mls.empty:
        filename = f"{output_prefix}_missing_in_MLS.xlsx"
        df_output = df_missing_mls.copy()
        
        # Save to Excel first
        df_output.to_excel(filename, index=False, sheet_name='Missing in MLS', engine='openpyxl')
        
        # Add hyperlinks using openpyxl
        if 'Parcel_ID' in df_output.columns and PARCEL_ID_URL_TEMPLATE:
            wb = load_workbook(filename)
            ws = wb['Missing in MLS']
            
            parcel_col_idx = list(df_output.columns).index('Parcel_ID') + 1
            
            for row_idx in range(2, len(df_output) + 2):
                cell = ws.cell(row=row_idx, column=parcel_col_idx)
                parcel_value = cell.value
                if parcel_value and str(parcel_value).strip():
                    url = PARCEL_ID_URL_TEMPLATE.format(parcel_id=parcel_value)
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'
            
            wb.save(filename)
        
        print(f"✓ Missing in MLS report saved: {filename} ({len(df_missing_mls)} records)")
        reports_generated.append(filename)

    if not df_value_mismatches.empty:
        filename = f"{output_prefix}_value_mismatches.xlsx"
        df_output = df_value_mismatches.copy()
        
        # Save to Excel first
        df_output.to_excel(filename, index=False, sheet_name='Value Mismatches', engine='openpyxl')
        
        # Add hyperlinks using openpyxl
        wb = load_workbook(filename)
        ws = wb['Value Mismatches']
        
        # Add Parcel_ID hyperlinks
        if 'Parcel_ID' in df_output.columns and PARCEL_ID_URL_TEMPLATE:
            parcel_col_idx = list(df_output.columns).index('Parcel_ID') + 1
            
            for row_idx in range(2, len(df_output) + 2):
                cell = ws.cell(row=row_idx, column=parcel_col_idx)
                parcel_value = cell.value
                if parcel_value and str(parcel_value).strip():
                    url = PARCEL_ID_URL_TEMPLATE.format(parcel_id=parcel_value)
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'
        
        # Add Zillow hyperlinks to Address column
        if 'Address' in df_output.columns:
            address_col_idx = list(df_output.columns).index('Address') + 1
            
            # Check if city and zip columns exist
            has_location_data = all(col in df_output.columns for col in ['City', 'Zip'])
            
            if has_location_data:
                city_col_idx = list(df_output.columns).index('City') + 1
                zip_col_idx = list(df_output.columns).index('Zip') + 1
                
                for row_idx in range(2, len(df_output) + 2):
                    address_cell = ws.cell(row=row_idx, column=address_col_idx)
                    address_value = address_cell.value
                    
                    if address_value and str(address_value).strip():
                        # Get city and zip from the same row
                        city = ws.cell(row=row_idx, column=city_col_idx).value
                        zip_code = ws.cell(row=row_idx, column=zip_col_idx).value
                        
                        url = format_zillow_url(address_value, city, 'OH', zip_code)
                        if url:
                            address_cell.hyperlink = url
                            address_cell.style = 'Hyperlink'
        
        wb.save(filename)
        
        print(f"✓ Value Mismatches report saved: {filename} ({len(df_value_mismatches)} mismatches)")
        reports_generated.append(filename)

    if not df_perfect_matches.empty:
        filename = f"{output_prefix}_perfect_matches.xlsx"
        df_output = df_perfect_matches.copy()
        
        # Save to Excel first
        df_output.to_excel(filename, index=False, sheet_name='Perfect Matches', engine='openpyxl')
        
        # Add hyperlinks using openpyxl
        wb = load_workbook(filename)
        ws = wb['Perfect Matches']
        
        # Add Parcel_ID hyperlinks
        if 'Parcel_ID' in df_output.columns and PARCEL_ID_URL_TEMPLATE:
            parcel_col_idx = list(df_output.columns).index('Parcel_ID') + 1
            
            for row_idx in range(2, len(df_output) + 2):
                cell = ws.cell(row=row_idx, column=parcel_col_idx)
                parcel_value = cell.value
                if parcel_value and str(parcel_value).strip():
                    url = PARCEL_ID_URL_TEMPLATE.format(parcel_id=parcel_value)
                    cell.hyperlink = url
                    cell.style = 'Hyperlink'
        
        # Add Zillow hyperlinks to Address column
        if 'Address' in df_output.columns:
            address_col_idx = list(df_output.columns).index('Address') + 1
            
            # Check if city and zip columns exist
            has_location_data = all(col in df_output.columns for col in ['City', 'Zip'])
            
            if has_location_data:
                city_col_idx = list(df_output.columns).index('City') + 1
                zip_col_idx = list(df_output.columns).index('Zip') + 1
                
                for row_idx in range(2, len(df_output) + 2):
                    address_cell = ws.cell(row=row_idx, column=address_col_idx)
                    address_value = address_cell.value
                    
                    if address_value and str(address_value).strip():
                        # Get city and zip from the same row
                        city = ws.cell(row=row_idx, column=city_col_idx).value
                        zip_code = ws.cell(row=row_idx, column=zip_col_idx).value
                        
                        url = format_zillow_url(address_value, city, 'OH', zip_code)
                        if url:
                            address_cell.hyperlink = url
                            address_cell.style = 'Hyperlink'
        
        wb.save(filename)
        
        print(f"✓ Perfect Matches report saved: {filename} ({len(df_perfect_matches)} records)")
        reports_generated.append(filename)

    if not reports_generated:
        print("\nNo discrepancies found - no reports generated.")

    return reports_generated

# --- Main Execution ---

if __name__ == "__main__":
    print("="*80)
    print("MLS vs. CAMA Data Comparison - Enhanced Version with Categorical Comparison")
    print("="*80)

    DEBUG_MODE = False  # Set to True to see detailed comparison info
    RUN_DIAGNOSTICS = False  # Set to True to run diagnostic analysis

    # 1. Load data
    mls_data = read_mls_data(MLS_DATA_PATH)
    cama_data = read_cama_data(CAMA_DATA_PATH)

    if mls_data is not None and cama_data is not None:
        mls_id_col_name = UNIQUE_ID_COLUMN.get('mls_col')
        cama_id_col_name = UNIQUE_ID_COLUMN.get('cama_col')

        if mls_id_col_name is None or cama_id_col_name is None:
            print("Error: UNIQUE_ID_COLUMN dictionary is missing required keys.")
        elif mls_id_col_name not in mls_data.columns:
            print(f"Error: Unique ID column '{mls_id_col_name}' not found in MLS data.")
        elif cama_id_col_name not in cama_data.columns:
            print(f"Error: Unique ID column '{cama_id_col_name}' not found in CAMA data.")
        else:
            print(f"\n📊 Data Summary:")
            print(f"   MLS records: {len(mls_data)}")
            print(f"   CAMA records: {len(cama_data)}")
            print(f"   Numeric tolerance: {NUMERIC_TOLERANCE}")

            # 2. Check for duplicates
            print("\n" + "="*80)
            print("STEP 1: Checking for Duplicate IDs")
            print("="*80)
            find_duplicate_ids(mls_data, mls_id_col_name, "MLS")
            find_duplicate_ids(cama_data, cama_id_col_name, "CAMA")

            # 3. Compare data
            print("\n" + "="*80)
            print("STEP 2: Comparing Data")
            print("="*80)

            df_missing_cama, df_missing_mls, df_value_mismatches, matched_records, df_perfect_matches = \
                compare_data_enhanced(mls_data, cama_data, UNIQUE_ID_COLUMN,
                                     COLUMNS_TO_COMPARE, cols_to_compare_sum=COLUMNS_TO_COMPARE_SUM,
                                     cols_to_compare_categorical=COLUMNS_TO_COMPARE_CATEGORICAL,
                                     debug_mode=DEBUG_MODE)

            # 4. Display results
            print("\n" + "="*80)
            print("STEP 3: Results Summary")
            print("="*80)

            print(f"\n✓ Records matched on {cama_id_col_name}: {len(matched_records)}")
            print(f"✗ Records missing in CAMA: {len(df_missing_cama)}")
            print(f"✗ Records missing in MLS: {len(df_missing_mls)}")
            print(f"⚠ Value mismatches found: {len(df_value_mismatches)}")
            print(f"✅ Perfect matches found: {len(df_perfect_matches)}")

            if not df_value_mismatches.empty:
                print("\n📊 Mismatches by Field:")
                mismatch_counts = df_value_mismatches['Field_MLS'].value_counts()
                for field, count in mismatch_counts.items():
                    print(f"   {field}: {count} mismatches")

            # 5. Generate reports
            print("\n" + "="*80)
            print("STEP 4: Generating CSV Reports")
            print("="*80)
            report_discrepancies_enhanced(df_missing_cama, df_missing_mls,
                                         df_value_mismatches, df_perfect_matches)

    else:
        print("❌ Data loading failed. Please check file paths and formats.")

    print("\n" + "="*80)
    print("Automation Complete")
    print("="*80)

    print("\n" + "="*80)
    print("Script Complete")
    print("="*80)

    # Download all reports if running in Colab
    try:
        from google.colab import files
        print("\n📥 Downloading Excel reports...")
        files.download('discrepancies_missing_in_CAMA.xlsx')
        files.download('discrepancies_missing_in_MLS.xlsx')
        files.download('discrepancies_value_mismatches.xlsx')
        files.download('discrepancies_perfect_matches.xlsx')
        print("✓ All reports downloaded!")
    except:
        print("\n💡 Files saved locally. Check your folder for the reports.")
